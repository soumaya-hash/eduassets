from collections import Counter
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from factures.models import Academie, Commune, DirectionProvinciale, Etablissement


REQUIRED_HEADERS = {
    'CD_PROV',
    'CD_MIL',
    'cd_com',
    'la_com',
    'll_com',
    'CD_ETAB',
    'NOM_ETABL',
    'NOM_ETABA',
    'typeEtab',
    'CD_etabr',
}


def normalize(value):
    """Convert spreadsheet cells to stable, comparable source codes."""
    if value is None:
        return ''
    return str(value).strip().upper()


def text(value):
    return '' if value is None else str(value).strip()


class Command(BaseCommand):
    help = 'Importe les établissements d’un fichier Excel source, de manière idempotente.'

    def add_arguments(self, parser):
        parser.add_argument('--fichier', required=True, help='Chemin du fichier Excel (.xlsx).')
        parser.add_argument('--academie-code', required=True)
        parser.add_argument('--academie-nom', required=True)
        parser.add_argument('--dp-code', default='227')
        parser.add_argument('--dp-nom', default='Province de Fahs-Anjra')
        parser.add_argument(
            '--appliquer',
            action='store_true',
            help='Écrit réellement les données. Sans cette option, seule la validation est effectuée.',
        )

    def handle(self, *args, **options):
        path = Path(options['fichier'])
        if not path.is_file():
            raise CommandError(f'Fichier introuvable : {path}')
        if path.suffix.lower() != '.xlsx':
            raise CommandError('Le fichier doit être au format .xlsx.')

        rows = self._read_rows(path)
        self._validate_province(rows, normalize(options['dp_code']))

        summary = self._summarize(rows)
        self._print_summary(summary, apply=options['appliquer'])
        if not options['appliquer']:
            self.stdout.write(self.style.WARNING('Validation terminée : aucune donnée n’a été écrite.'))
            return

        with transaction.atomic():
            result = self._import_rows(rows, options)

        self.stdout.write(self.style.SUCCESS(
            'Import terminé : '
            f"{result['communes_created']} commune(s) créée(s), "
            f"{result['etablissements_created']} établissement(s) créé(s), "
            f"{result['etablissements_updated']} établissement(s) mis à jour, "
            f"{result['parents_linked']} rattachement(s) établi(s)."
        ))
        if result['unresolved_parent_codes']:
            self.stdout.write(self.style.WARNING(
                'Codes d’établissement principal absents du fichier : '
                + ', '.join(sorted(result['unresolved_parent_codes']))
            ))

    def _read_rows(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = next(values, None)
            if not headers:
                raise CommandError('Le fichier Excel est vide.')

            normalized_headers = [text(header) for header in headers]
            missing_headers = REQUIRED_HEADERS - set(normalized_headers)
            if missing_headers:
                raise CommandError('Colonnes manquantes : ' + ', '.join(sorted(missing_headers)))

            rows = []
            for line_number, values_row in enumerate(values, start=2):
                row = dict(zip(normalized_headers, values_row))
                if not any(value not in (None, '') for value in row.values()):
                    continue
                if not normalize(row['CD_ETAB']):
                    raise CommandError(f'Ligne {line_number} : CD_ETAB est obligatoire.')
                if not text(row['NOM_ETABL']):
                    raise CommandError(f'Ligne {line_number} : NOM_ETABL est obligatoire.')
                rows.append(row)
            return rows
        finally:
            workbook.close()

    def _validate_province(self, rows, expected_code):
        province_codes = {normalize(row['CD_PROV']) for row in rows}
        if province_codes != {expected_code}:
            raise CommandError(
                f"Le fichier contient les provinces {', '.join(sorted(province_codes))}; "
                f'la DP attendue est {expected_code}.'
            )

    def _summarize(self, rows):
        codes = [normalize(row['CD_ETAB']) for row in rows]
        duplicates = [code for code, count in Counter(codes).items() if count > 1]
        if duplicates:
            raise CommandError('Codes établissement dupliqués : ' + ', '.join(sorted(duplicates)))
        return {
            'records': len(rows),
            'communes': len({normalize(row['cd_com']) for row in rows}),
            'types': Counter(text(row['typeEtab']) for row in rows),
            'parent_links': sum(1 for row in rows if normalize(row['CD_etabr'])),
        }

    def _print_summary(self, summary, apply):
        mode = 'IMPORT' if apply else 'VALIDATION'
        self.stdout.write(f'{mode} — {summary["records"]} établissements, {summary["communes"]} communes.')
        self.stdout.write('Répartition par type : ' + ', '.join(
            f'{key}: {value}' for key, value in sorted(summary['types'].items())
        ))
        self.stdout.write(f"Rattachements d’établissement principal à traiter : {summary['parent_links']}.")

    def _import_rows(self, rows, options):
        academie, _ = Academie.objects.get_or_create(
            code=normalize(options['academie_code']),
            defaults={'nom': text(options['academie_nom'])},
        )
        if academie.nom != text(options['academie_nom']):
            academie.nom = text(options['academie_nom'])
            academie.save(update_fields=['nom'])

        dp, _ = DirectionProvinciale.objects.get_or_create(
            code=normalize(options['dp_code']),
            defaults={'nom': text(options['dp_nom']), 'academie': academie},
        )
        dp_updated_fields = []
        if dp.nom != text(options['dp_nom']):
            dp.nom = text(options['dp_nom'])
            dp_updated_fields.append('nom')
        if dp.academie_id != academie.id:
            dp.academie = academie
            dp_updated_fields.append('academie')
        if dp_updated_fields:
            dp.save(update_fields=dp_updated_fields)

        communes = {}
        communes_created = 0
        for row in rows:
            commune_code = normalize(row['cd_com'])
            if commune_code in communes:
                continue
            commune_name = text(row['ll_com'])
            commune = Commune.objects.filter(
                code=commune_code,
                direction_provinciale=dp,
            ).first()
            if commune is None:
                commune = Commune.objects.filter(
                    nom=commune_name,
                    direction_provinciale=dp,
                ).first()
            created = commune is None
            if created:
                commune = Commune.objects.create(
                    code=commune_code,
                    nom=commune_name,
                    direction_provinciale=dp,
                )
            elif commune.code != commune_code:
                commune.code = commune_code
                commune.save(update_fields=['code'])
            if commune.nom != commune_name:
                commune.nom = commune_name
                commune.save(update_fields=['nom'])
            communes[commune_code] = commune
            communes_created += int(created)

        establishments = {}
        establishments_created = 0
        establishments_updated = 0
        for row in rows:
            code = normalize(row['CD_ETAB'])
            defaults = {
                'nom': text(row['NOM_ETABL']),
                'nom_arabe': text(row['NOM_ETABA']) or None,
                'type_etablissement': text(row['typeEtab']) or None,
                'code_milieu': normalize(row['CD_MIL']) or None,
                'academie': academie,
                'direction_provinciale': dp,
                'commune': communes[normalize(row['cd_com'])],
                'actif': True,
            }
            establishment, created = Etablissement.objects.update_or_create(
                identifiant_unique=code,
                defaults=defaults,
            )
            establishments[code] = establishment
            establishments_created += int(created)
            establishments_updated += int(not created)

        parents_linked = 0
        unresolved_parent_codes = set()
        for row in rows:
            child = establishments[normalize(row['CD_ETAB'])]
            parent_code = normalize(row['CD_etabr'])
            parent = establishments.get(parent_code)
            if parent_code and parent is None:
                unresolved_parent_codes.add(parent_code)
                continue
            if child.etablissement_rattachement_id != (parent.id if parent else None):
                child.etablissement_rattachement = parent
                child.save(update_fields=['etablissement_rattachement'])
            if parent:
                parents_linked += 1

        return {
            'communes_created': communes_created,
            'etablissements_created': establishments_created,
            'etablissements_updated': establishments_updated,
            'parents_linked': parents_linked,
            'unresolved_parent_codes': unresolved_parent_codes,
        }
